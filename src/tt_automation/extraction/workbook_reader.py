from datetime import date, datetime
from io import BytesIO
import json
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlrd

from tt_automation.extraction.documents import SourceDocument


MAX_SHEETS = 10
MAX_ROWS_PER_SHEET = 300
MAX_COLUMNS_PER_SHEET = 50
MAX_WORKBOOK_CHARACTERS = 50_000


class WorkbookReadError(ValueError):
    """Raised when an uploaded workbook cannot be normalized for extraction."""


def workbook_to_text(document: SourceDocument) -> str:
    """Render non-empty workbook cells as compact, position-aware text."""

    if not document.is_workbook:
        raise WorkbookReadError(f"{document.name} is not a supported workbook.")

    try:
        if document.suffix == ".xls":
            text = _read_legacy_workbook(document.content)
        else:
            text = _read_modern_workbook(document.content)
    except Exception as error:
        raise WorkbookReadError(f"Could not read workbook: {document.name}") from error

    if not text.strip():
        raise WorkbookReadError(
            f"Workbook contains no readable values: {document.name}"
        )
    if len(text) > MAX_WORKBOOK_CHARACTERS:
        return f"{text[:MAX_WORKBOOK_CHARACTERS]}\n[Workbook content truncated]"
    return text


def _read_modern_workbook(content: bytes) -> str:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    lines: list[str] = []

    try:
        for sheet in workbook.worksheets[:MAX_SHEETS]:
            lines.append(f"[Sheet: {sheet.title}]")
            maximum_row = min(sheet.max_row, MAX_ROWS_PER_SHEET)
            maximum_column = min(sheet.max_column, MAX_COLUMNS_PER_SHEET)
            rows = sheet.iter_rows(
                min_row=1,
                max_row=maximum_row,
                max_col=maximum_column,
                values_only=True,
            )
            for row_number, values in enumerate(rows, start=1):
                if line := _render_row(row_number, values):
                    lines.append(line)
            if sheet.max_row > maximum_row or sheet.max_column > maximum_column:
                lines.append("[Sheet content truncated]")
    finally:
        workbook.close()

    if len(workbook.sheetnames) > MAX_SHEETS:
        lines.append("[Additional sheets omitted]")
    return "\n".join(lines)


def _read_legacy_workbook(content: bytes) -> str:
    """Read a legacy .xls workbook and render its cells as a positoin-aware text.
    :param content: Raw bytes of the legacy .xls workbook.
    : returns: Newline-separated text where each non-empty cell within the configured
        row and column limits is represented by its Excel row and column.
        Adds truncation markers when applicable.
    """
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    lines: list[str] = []

    try:
        for sheet in workbook.sheets()[:MAX_SHEETS]:
            lines.append(f"[Sheet: {sheet.name}]")
            maximum_row = min(sheet.nrows, MAX_ROWS_PER_SHEET)
            maximum_column = min(sheet.ncols, MAX_COLUMNS_PER_SHEET)
            for row_index in range(maximum_row):
                values = [
                    _legacy_cell_value(workbook, sheet.cell(row_index, column_index))
                    for column_index in range(maximum_column)
                ]
                if line := _render_row(row_index + 1, values):
                    lines.append(line)
            if sheet.nrows > maximum_row or sheet.ncols > maximum_column:
                lines.append("[Sheet content truncated]")
    finally:
        workbook.release_resources()

    if workbook.nsheets > MAX_SHEETS:
        lines.append("[Additional sheets omitted]")
    return "\n".join(lines)


def _legacy_cell_value(workbook: xlrd.book.Book, cell: xlrd.sheet.Cell) -> Any:
    """Normalize cell values from the legacy xlrd library to match openpyxl output.
    :params workbook: The xlrd workbook object, needed for date conversion.
    :params cell: The xlrd cell object to normalize.
    :returns: The normalized cell value, which may be a string, number, boolean, or datetime.
    """
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    return cell.value


def _render_row(row_number: int, values: tuple[Any, ...] | list[Any]) -> str:
    rendered_cells = []
    for column_number, value in enumerate(values, start=1):
        if value is None or value == "":
            continue
        column = get_column_letter(column_number)
        rendered_cells.append(f"{column}={_render_value(value)}")

    if not rendered_cells:
        return ""
    return f"Row {row_number}: {' | '.join(rendered_cells)}"


def _render_value(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        value = value.isoformat()
    return json.dumps(value, ensure_ascii=True, default=str)
